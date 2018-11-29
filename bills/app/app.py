from typing import List, Dict
from flask import Flask, request, send_from_directory
import mysql.connector
import openpyxl as xl
import json
import requests
import logging
import datetime
import re # Regular Expression
from datetime import datetime
from pathlib import Path


app = Flask(__name__, static_url_path='/')


logging.basicConfig(filename = 'billingserviceapp.log', level = logging.DEBUG, format = '%(asctime)s:%(levelname)s:%(funcName)s:%(message)s')

databaseConfig = {
        'user': 'root',
        'password': 'root',
        'host': 'billingservicedb',
        'port': '3306',
        'database': 'flaskApp'
    }

@app.route('/truck', methods=["POST"])
def truckInsert():
    try:
       connection = mysql.connector.connect(**databaseConfig)
       cursor = connection.cursor()
       cursor.execute('SELECT * FROM provider WHERE providerId = %d'%(int(request.form["provider"])))
       myProviderId = [{providerId} for (providerId) in cursor]
       if myProviderId:
          cursor.execute('INSERT INTO truck VALUES (%d, %d)'%((int(request.form["id"])),int(request.form["provider"])))
          connection.commit()
          cursor.close()
          connection.close()
          return "ok"
       else:
         logging.error("This provider does not exist in the system")
         return ("This %d provider does not exist in the system"%(int(request.form["provider"])))
    except Exception as e:
        logging.error("Failed to add %d provider"%(int(request.form["id"])))
        return str(e)
 
@app.route('/provider', methods=["POST"])
def providerInsert():
       logging.info('Adding new provider to the table')
       try:
        connection = mysql.connector.connect(**databaseConfig)
        cursor = connection.cursor()
        cursor.execute('INSERT INTO provider VALUES (NULL, "%s")'%(request.form["name"]))
        connection.commit()
        providerId = cursor.lastrowid
        cursor.close()
        response = { "id": providerId}
        connection.close()
        logging.info('A %s provider was added successfully'%(request.form["name"]))
        return str(response)
       except Exception as e:
          logging.error("Failed to add provider %s"%(request.form["name"]))
          return str(e)

@app.route('/truckList')
def listTruck() -> List[Dict]:
    try:
     connection = mysql.connector.connect(**databaseConfig)
     cursor = connection.cursor()
     cursor.execute('SELECT * FROM truck')
     results = [{truckId: providerId} for (truckId, providerId) in cursor]
     cursor.close()
     connection.close()
     logging.info('Show all trucks successfully completed')
     return str(results)
    except Exception as e:
        logging.error("Failed to view all trucks")
        return str(e)

@app.route('/getrates')
def getRates():
    try:
        return send_from_directory('in', "rates.xlsx")
        logging.info("rates.xlsx file downloaded successfully")
    except Exception as e:
        logging.error("Failed to download file rates.xlsx")
        return str(e)


@app.route('/providerList')
def providerList() -> List[Dict]:
    try:
     connection = mysql.connector.connect(**databaseConfig)
     cursor = connection.cursor()
     cursor.execute('SELECT * FROM provider')
     results = [{providerId: providerName} for (providerId, providerName) in cursor]
     cursor.close()
     connection.close()
     logging.info('Show all providers successfully completed')
     return str(results)
    except Exception as e:
        logging.error("Failed to view all providers")
        return str(e)


@app.route('/provider/<id>', methods=["POST"])
def providerUpdate(id):
    try:
        connection = mysql.connector.connect(**databaseConfig)
        cursor = connection.cursor()  
        cursor.execute('UPDATE provider SET providerName = "{0}" WHERE providerId = {1}'.format(request.form["providerName"], id))
        connection.commit()
        cursor.close()
        connection.close()
        logging.info('%s provider Update'%id)
        return "ok"
    except Exception as e:
        logging.error("Failed to update %s provider"%id)
        return str(e)
    


@app.route('/tests', methods=["GET"])
def tests():
    try:
        response = requests.get('http://localhost:6002/health')
        return str(response)
    except Exception as e:
        return str(e)


@app.route('/truck/<id>', methods=["GET"])
def get_truck(id):

    date_from = request.args.get("from")
    date_to = request.args.get("to")

    if id is None:
        logging.info("No truck id")
        return "No truck id"
    elif date_from is None:
        # By default the date_from is from the first day of the current month
        date_from = datetime.now().strftime('%Y%m01000000')
    elif re.match("^[0-9]{14}$", date_from) is None:
        logging.error("The format of date from {0} is not correct".format(date_from))
        return str ("The format of date from {0} is not correct".format(date_from))

    # Checking the dateTo format
    if date_to is None:
        # By default the date_to is now
        date_to = datetime.now().strftime('%Y%m%d%H%M%S')
    elif re.match("^[0-9]{14}$", date_to) is None:
        logging.error("The format of date to {0} is not correct".format(date_to))
        return str("The format of date to {0} is not correct".format(date_to))
    try:
        response = requests.get('http://service_app_weight:5000/item/{0}?from={1}&to={2}'.format(id,date_from,date_to))

        if response.status_code == 404:
            logging.error("Truck not found")
            return "Truck not found"

        logging.info("successfully get truck id={0} from={1} to={2}".format(id,date_from,date_to))
        return response.json()
    except Exception as error:
        logging.error("get_truck: {0}".format( error))
        return str("get_truck: {0}".format( error))



@app.route('/truckList')
def truckList():
    try:
        connection = mysql.connector.connect(**databaseConfig)
        cursor = connection.cursor()
        cursor.execute('SELECT * FROM truck')
        results = [{truckId: providerId} for (truckId, providerId) in cursor]
        cursor.close()
        connection.close()
        logging.info('Show all trucks successfully completed')
        return str(results)
    except Exception as e:
        logging.error("Failed to view all trucks")
        return str(e)



@app.route('/productList')
def productList():
    connection = mysql.connector.connect(**databaseConfig)
    cursor = connection.cursor()
    cursor.execute('SELECT * FROM rates')
    results = cursor.fetchall()
    cursor.close()
    connection.close()
    return str(results)

    
@app.route('/truck/<id>', methods=["POST"])
def truckUpdate(id):
    try:
        connection = mysql.connector.connect(**databaseConfig)
        cursor = connection.cursor()  
        cursor.execute('UPDATE truck SET providerId = "{0}" WHERE truckId = {1}'.format(request.form["providerId"], id))
        connection.commit()
        cursor.close()
        connection.close()
        logging.info('%s truck Update'%id)
        return "ok"
    except Exception as e:
        logging.error("Failed to update %s provider"%id)
        return str(e)
    

@app.route("/rates",methods=["POST"])
def postrates():
    filename = request.args.get("file")

    try:
        wb = xl.load_workbook("/in/" + filename )
        ws = wb.get_active_sheet()
        connection = mysql.connector.connect(**databaseConfig)
        cursor = connection.cursor()
        sql_insert_rates_query = "INSERT INTO rates (productName, scope, rates) VALUES (%s, %s, %s)"
        cursor.execute('TRUNCATE TABLE rates')
        row = 2
        while ws.cell(row, 1).value is not None:
            productName = ws.cell(row, 1).value
            rate = ws.cell(row, 2).value
            scope = ws.cell(row, 3).value
            insert_tuple = (productName, scope, rate)
            cursor.execute(sql_insert_rates_query, insert_tuple)
            row += 1

        connection.commit()
        cursor.close()
        connection.close()
        logging.info("RATES UPLOADED")
        return "RATES UPLOADED"
    except FileNotFoundError:
        logging.error("File Not Found")
        return "File Not Found"

    except mysql.connector.Error as error:
        logging.error("Rates uploading failed {}".format(error))
        return "Rates uploading failed {}".format(error)

    except Exception as error:
        logging.error("Error {}".format(error))
        return "Error {}".format(error)



@app.route('/bill/<id>')
def getBills(id):
    providerId = id
    dateFrom = request.args.get("from")
    dateTo = request.args.get("to")
    productsIdList = []
    trucksIdList = []
    products = {}
    data = {}
    data["id"] = providerId
    data["truckCount"] = 0
    data["sessionCount"] = 0
    data["products"] = 0
    data["total"] = 0
    # Checking the dateFrom format
    if dateFrom is None:
        # By default the dateFrom is from the first day of the current month
        dateFrom = datetime.datetime.now().strftime('%Y%m01000000')
    elif re.match("^[0-9]{14}$", dateFrom) is None :
        logging.error("/bill/<id> : The format of from Date is not correct")
        return "The format of from Date is not correct"

    # Checking the dateTo format
    if dateTo is None:
        # By default the dateTo now
        dateTo = datetime.datetime.now().strftime('%Y%m%d%H%M%S')
    elif re.match("^[0-9]{14}$", dateTo) is None :
        logging.error("/bill/<id> : The format of to Date is not correct")
        return "The format of to Date is not correct"
    
    data["from"] = dateFrom
    data["to"] = dateTo
    # Oppening the connection 
    connection = mysql.connector.connect(**databaseConfig)
    
    # Getting provider infos
    cursor = connection.cursor()
    cursor.execute('SELECT * FROM provider WHERE providerId = "{0}"'.format(providerId))
    provider = cursor.fetchall()
    cursor.close()
    if provider != []:
        providerName = provider[0][1]
        data["name"] = providerName
    else:
        # Don't forget to close the connection 
        connection.close()
        logging.error("/bill/<id> : Error while fetching the provider details")
        return "Error while fetching the provider details"

    # Getting all the truck of the providerId
    cursor = connection.cursor()
    cursor.execute('SELECT DISTINCT(truckId) FROM truck WHERE providerId = "{0}"'.format(providerId))
    truckList = cursor.fetchall()
    cursor.close()
    #return str(truckList)
    if truckList != []:
        for truckId in truckList:
            try:
                #r_weight_itemfunction = requests.get('http://service_app_weight:5000/item/{0}?from={1}&to={2}'.format(truckId,dateFrom,dateTo))
                #itemReturn = r_weight_item.json()
                itemReturn = '{ "id": 1,"tara": 10,"sessions": [ 1,2,3 ] }'
                #itemReturn = '{}'
                if itemReturn != "{}":
                    #return itemReturn
                    itemReturnAsTable = json.loads(itemReturn)
                    sessionList = itemReturnAsTable["sessions"]
                    currentTruckTara = itemReturnAsTable["tara"]
                    currentTruckId = itemReturnAsTable["id"]
                    if currentTruckId not in trucksIdList:
                        data["truckCount"] += 1
                        trucksIdList.append(currentTruckId)
                    for sessionId in sessionList:
                        #r_weight_weightfunction = requests.get('http://service_app_weight:5000/weight/{0}?from={1}&to={2}&filter=out'.format(truckId,dateFrom,dateTo))
                        #sessionContentReturn = r_weight_weightfunction.json()
                        sessionsContentReturn = '[{ "id": 1,"direction": "out","bruto": 100,"neto": 20,"produce": "product1","containers": [ 1,2 ]},{ "id": 3,"direction": "in","bruto": 200,"neto": 40,"produce": "product2","containers": [ 1,2 ]},{ "id": 5,"direction": "out","bruto": 100,"neto": 20,"produce": "product1","containers": [ 1,2 ]},{ "id": 2,"direction": "out","bruto": 100,"neto": 20,"produce": "product1","containers": [ 1,2 ]}]'
                        #sessionsContentReturn = '[{ "id": 2,"direction": "out","bruto": 100,"neto": 20,"produce": "product1","containers": [ 1,2 ]}]'
                        if sessionsContentReturn != "{}":
                            sessionsContentReturnAsTable = json.loads(sessionsContentReturn)
                            for oneSession in sessionsContentReturnAsTable:
                                # check if the session id is in our session table
                                #return oneSession["id"]
                                if oneSession["id"] in sessionList and oneSession["neto"] != "na":
                                    productName = oneSession["produce"]
                                    if oneSession["produce"] in productsIdList:
                                        data["sessionCount"] += 1
                                        products[productName]["count"] += 1
                                        products[productName]["amount"] += int(oneSession["neto"])
                                        tempPay = products[productName]["pay"]
                                        products[productName]["pay"] = products[productName]["amount"] * products[productName]["rate"]
                                    else:
                                        # Fetching the product rate corresponding to this provider
                                        cursor = connection.cursor()
                                        cursor.execute('SELECT * FROM rates WHERE productName = "{0}" AND (scope = "{1}" OR scope = "ALL") ORDER BY scope'.format(productName, providerId))
                                        productRateForThisProvider = cursor.fetchall()
                                        cursor.close()
                                        if productRateForThisProvider != []:
                                            data["sessionCount"] += 1
                                            products[productName] = {}
                                            products[productName]["rate"] = int(productRateForThisProvider[0][2])
                                            products[productName]["count"] = 1
                                            products[productName]["amount"] = int(oneSession["neto"])
                                            products[productName]["pay"] = products[productName]["amount"] * products[productName]["rate"]
                                        else: 
                                            continue
                                else: 
                                    continue
                        else:
                            logging.error("/bill/<id> : Impossible to fetch the session content")
                            # Don't forget to close the connection 
                            connection.close()
                            return "Impossible to fetch the session content"
                else:
                    # if the itemReturn is empty so we don't have to do any calculation
                    logging.info("/bill/<id> : Successfull call")
                    # Don't forget to close the connection 
                    connection.close()
                    return json.dumps(data)
                #return str(products)
                if products != {}:
                    dataProducts = []
                    for product in products:
                        #return  str(products[product]["amount"])
                        tempProduct = {}
                        tempProduct["product"] = product
                        tempProduct["count"] = products[product]["count"]
                        tempProduct["amount"] = products[product]["amount"]
                        tempProduct["rate"] = products[product]["rate"]
                        tempProduct["pay"] = products[product]["pay"]
                        data["total"] += products[product]["pay"]
                        dataProducts.append(tempProduct)
                    data["products"] = dataProducts
                logging.info("/bill/<id> : Successfull call")
                # Don't forget to close the connection 
                connection.close()
                return json.dumps(data)
            except Exception as e:
                # Don't forget to close the connection 
                connection.close()
                logging.error("/bill/<id> : Error {}".format(e))
                return str(e)
    else:
        logging.error("/bill/<id> : The truck list is empty for this provider")
        # Don't forget to close the connection 
        connection.close()
        return "The truck list is empty for this provider"
    # Don't forget to close the connection 
    connection.close()
    return "END OF THE FUNCTION"




@app.route('/')
def index() -> str:
    print("HI EVERY BODY")
    return "IndexPage"


@app.route('/health')
def health():
    # test db connection
    try:
        cnx = mysql.connector.connect(**databaseConfig)
        cnx.close()
    except Exception as e:
        logging.error('Database connection failed with error %s' % e)
        return 'Error: %s' % e

    # test existence of /in dir
    try:
        path = '../in'
        if isdir(path) and islink(path):
            pass
    except Exception as e:
        logging.error('`/in` directory doesn\'t exist.')
        return 'Error: %s' % e

    return 'ok'

if __name__ == '__main__':
    logging.info('Starting Flask server...')
    print("Hi Bro")
    app.run(host='0.0.0.0',debug=True)

